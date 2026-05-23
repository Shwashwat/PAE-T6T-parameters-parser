# default values
# CAUTION!!!! 'File Name' is hard coded if you want file name
monthly_params = ['File Name', 'Frequency (MHz)' , 'Fwd Power' , 'Rfl Power' , 
                  'Line In Level (dBm)' , 'Inhibit' , 'Transmit Timeout (Sec)' ,
                  'Mod Depth (%)' , 'TX VOGAD' , 'Key Priority' , 'Ready Output Polarity',
                  'PTT Input Polarity' , 'Phantom PTT Input Polarity' ,
                  'PTT Reference Voltage (V)' , 'PTT Output Polarity',
                  ]

from os import listdir, path
from argparse import ArgumentParser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from itertools import zip_longest

def merge_dicts(list_of_dicts : list[dict]) -> dict:
   all_keys = list(dict.fromkeys(key for d in list_of_dicts for key in d))
   n = len(list_of_dicts)

   # pre-fill every col with None 
   merged_dict = {key: [None]*n for key in all_keys}

   # iterate ONLY over paramets exist per file
   for i, dictionary in enumerate(list_of_dicts):
      for key, value in dictionary.items():
         merged_dict[key][i] = value

   return merged_dict

# adds .xlsx extension if not already in the given name
def name_check(proposed_name:str) -> str :
   return ( 
      proposed_name 
      if proposed_name.endswith(".xlsx")
      else proposed_name + ".xlsx"
   )

def parse_cmd():
   # Create command-line parser
   parser = ArgumentParser(description = "A simple script to summarize PAE T6T settings")

   # Add arguments
   parser.add_argument(
      "-f", 
      "--folder", 
      required=True, 
      help="Folder where all txt are kept"
   )
   
   parser.add_argument(
      "-a", 
      "--all", 
      action="store_true", 
      help="Shows all the data collected"
   )

   parser.add_argument(
      "-p", 
      "--params",
      nargs="+", 
      default=monthly_params, 
      help="Parameters you need"
   )

   parser.add_argument(
      "-n", 
      "--save_file_name", 
      help="Report name"
   )

   parser.add_argument(
      "--extra_params",
      nargs="+",
      help="Add more parameters over the monthly ones"
   )

   return parser.parse_args()


def t6t_settings_parser(file_path:str, params_list:list[str])->dict: # file_path is the Path of file, paramslsit for checking
   
   settings = dict()
   with open(file_path, encoding="utf-8", errors="replace") as file:
      file_name = path.splitext(path.basename(file_path))[0] # -4 assumes 3 letter extension

      settings['File Name'] = file_name

      seen = {'File Name'} # add 'File Name' as this is in given params list

      duplicate_count_dict = dict()

      for raw_line in file:
         line = raw_line.strip()

         if not line: 
            continue

         if line.startswith("[") :

            # if it enounters this in any line rest of the file is skipped
            if "Bit History - " in line:
               break

            continue

         # to discard two weird hexadicimal numbers
         if line.startswith("0x"):
            continue

         if " : " not in line:
            continue

         try:
            name, value = line.split(" : ", 1)

            # process for duplicate keys
            if name not in seen: 
               settings[name] = value
               seen.add(name)
            else: 
               duplicate_count_dict[name] = duplicate_count_dict.get(name, 0) + 1
               duplicated_name = f"{name}__dup{duplicate_count_dict[name]}"
               settings[duplicated_name] = value
               seen.add(duplicated_name)
               print(f"[WARN] Parameter {name} appears {duplicate_count_dict[name]} times in current file")

         except ValueError:
            continue

         
      missing_cols = [col for col in params_list if col not in seen]

      if missing_cols:
         print(f"[WARN] Missing params: {missing_cols}")

      # return all_settings
   return settings

def export_to_formatted_excel(dict_of_lists:dict , output_path:str):
   wb = Workbook()
   bold = Font(bold=True)

   ws = wb.active # worksheet NOT in write only mode cuz formatting not avbl in write only mode

   headers = list(dict_of_lists.keys())

   # pre-compute widths
   max_widths = [len(h) for h  in headers]

   ws.append(headers)
   # convert columns to rows (in zip function)
   for row in zip_longest(*dict_of_lists.values()):
      ws.append(row)
      for i, value in enumerate(row):
         if value is not None:
            max_widths[i] = max(max_widths[i], len(str(value)))

   # apply column widths

   ### NOTE
   # tried differnet widths for first columns and smaller width for rest
   # resulting excel was nicer to read values from but many headers were similar
   # so needing manually extend it
   # RESULT: not worth implementing it
   ###
   
   MAX_COL_WIDTH = 20

   for i, width in enumerate(max_widths, start=1):
      ws.column_dimensions[get_column_letter(i)].width = min(width + 2, MAX_COL_WIDTH)

   # apply BOLD
   # header
   for cell in ws[1]:
      cell.font = bold

   # for first column
   for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
      row[0].font = bold

   wb.save(output_path)

def get_n_save_settings(folder:str, params:list[str] | None = None, extra_params=None, all_flag=False, save_file_name=None)-> tuple[dict , str]:

   # normalize the path delimiter
   folder = path.normpath(folder)
   
   # required params
   required_params = []
   if params is None:
      params = monthly_params.copy()
      
   required_params = params + extra_params if extra_params else params

   # read all contents of the folder
   files_in_dir = listdir(folder)

   # filter out unwanted files
   text_files = []

   for entry in files_in_dir:
      if path.isfile(path.join(folder, entry)) and entry.lower().endswith(".txt") :
         text_files.append(entry)

   # extract settings for each transmitter
   # master_param_dict = dict()

   # *** OPTIMIZATION: Use a list to store DataFrames, then concatenate once ***
   all_dictionaries = []

   # combine data from each file
   total = len(text_files)
   
   if total == 0:
      print("[WARN] No .txt files found to process.")
      return 
   
   print("[INFO] Processing started...\n")

   for i, text in enumerate(text_files, 1):
      print(f"[INFO] [{i}/{total}] Processing: {text}", flush=True)
      # Append the single-row DataFrame result to the list
      all_dictionaries.append(t6t_settings_parser( path.join(folder, text), required_params ))

   # Perform a single, much faster concatenation operation outside the loop
   master_param_dict = merge_dicts(all_dictionaries)

   # print(master_param_dict)
   # check if any save file name is provided# check if any save file name is provided
   if save_file_name:
      report_name = path.join(folder, save_file_name)
   else:
      # Fix: Extract only the folder name (e.g., 'FEB25') for the file_name
      folder_name = path.basename(folder)
      report_name = path.join(folder, folder_name + "_monthly.xlsx")

   dict_processed = master_param_dict if all_flag else {param:master_param_dict.get(param, []) for param in required_params}
   export_to_formatted_excel(dict_processed , name_check(report_name))

   # prints report details
   print("[INFO] Processing complete.\n")
   print(f"[INFO] Path processed: {folder}")
   print(f"[INFO_REPORT] REPORT SAVED AT: { report_name }")


   return dict_processed, report_name


## running code
# from sys import argv
# if len(argv):
if __name__ == "__main__":
    parameters = parse_cmd()
    get_n_save_settings(parameters.folder, parameters.params, parameters.extra_params, parameters.all, parameters.save_file_name)


# default values
# CAUTION!!!! 'File Name' is hard coded if you want file name
monthly_params = ['File Name', 'Frequency (MHz)' , 'Fwd Power' , 'Rfl Power' , 
                  'Line In Level (dBm)' , 'Inhibit' , 'Transmit Timeout (Sec)' ,
                  'Mod Depth (%)' , 'TX VOGAD' , 'Key Priority' , 'Ready Output Polarity',
                  'PTT Input Polarity' , 'Phantom PTT Input Polarity' ,
                  'PTT Reference Voltage (V)' , 'PTT Output Polarity',
                  ]

from os import listdir, path
from argparse import ArgumentParser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from itertools import zip_longest

def merge_dicts(list_of_dicts : list[dict]) -> dict:
   all_keys = list(dict.fromkeys(key for d in list_of_dicts for key in d))
   n = len(list_of_dicts)

   # pre-fill every col with None 
   merged_dict = {key: [None]*n for key in all_keys}

   # iterate ONLY over paramets exist per file
   for i, dictionary in enumerate(list_of_dicts):
      for key, value in dictionary.items():
         merged_dict[key][i] = value

   return merged_dict

# adds .xlsx extension if not already in the given name
def name_check(proposed_name:str) -> str :
   return ( 
      proposed_name 
      if proposed_name.endswith(".xlsx")
      else proposed_name + ".xlsx"
   )

def parse_cmd():
   # Create command-line parser
   parser = ArgumentParser(description = "A simple script to summarize PAE T6T settings")

   # Add arguments
   parser.add_argument(
      "-f", 
      "--folder", 
      required=True, 
      help="Folder where all txt are kept"
   )
   
   parser.add_argument(
      "-a", 
      "--all", 
      action="store_true", 
      help="Shows all the data collected"
   )

   parser.add_argument(
      "-p", 
      "--params",
      nargs="+", 
      default=monthly_params, 
      help="Parameters you need"
   )

   parser.add_argument(
      "-n", 
      "--save_file_name", 
      help="Report name"
   )

   parser.add_argument(
      "--extra_params",
      nargs="+",
      help="Add more parameters over the monthly ones"
   )

   return parser.parse_args()


def t6t_settings_parser(file_path:str, params_list:list[str])->dict: # file_path is the Path of file, paramslsit for checking
   
   settings = dict()
   with open(file_path, encoding="utf-8", errors="replace") as file:
      file_name = path.splitext(path.basename(file_path))[0] # -4 assumes 3 letter extension

      settings['File Name'] = file_name

      seen = {'File Name'} # add 'File Name' as this is in given params list

      duplicate_count_dict = dict()

      for raw_line in file:
         line = raw_line.strip()

         if not line: 
            continue

         if line.startswith("[") :

            # if it enounters this in any line rest of the file is skipped
            if "Bit History - " in line:
               break

            continue

         # to discard two weird hexadicimal numbers
         if line.startswith("0x"):
            continue

         if " : " not in line:
            continue

         try:
            name, value = line.split(" : ", 1)

            # process for duplicate keys
            if name not in seen: 
               settings[name] = value
               seen.add(name)
            else: 
               duplicate_count_dict[name] = duplicate_count_dict.get(name, 0) + 1
               duplicated_name = f"{name}__dup{duplicate_count_dict[name]}"
               settings[duplicated_name] = value
               seen.add(duplicated_name)
               print(f"[WARN] Parameter {name} appears {duplicate_count_dict[name]} times in current file")

         except ValueError:
            continue

         
      missing_cols = [col for col in params_list if col not in seen]

      if missing_cols:
         print(f"[WARN] Missing params: {missing_cols}")

      # return all_settings
   return settings

def export_to_formatted_excel(dict_of_lists:dict , output_path:str):
   wb = Workbook()
   bold = Font(bold=True)

   ws = wb.active # worksheet NOT in write only mode cuz formatting not avbl in write only mode

   headers = list(dict_of_lists.keys())

   # pre-compute widths
   max_widths = [len(h) for h  in headers]

   ws.append(headers)
   # convert columns to rows (in zip function)
   for row in zip_longest(*dict_of_lists.values()):
      ws.append(row)
      for i, value in enumerate(row):
         if value is not None:
            max_widths[i] = max(max_widths[i], len(str(value)))

   # apply column widths

   ### NOTE
   # tried differnet widths for first columns and smaller width for rest
   # resulting excel was nicer to read values from but many headers were similar
   # so needing manually extend it
   # RESULT: not worth implementing it
   ###
   
   MAX_COL_WIDTH = 20

   for i, width in enumerate(max_widths, start=1):
      ws.column_dimensions[get_column_letter(i)].width = min(width + 2, MAX_COL_WIDTH)

   # apply BOLD
   # header
   for cell in ws[1]:
      cell.font = bold

   # for first column
   for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
      row[0].font = bold

   wb.save(output_path)

def get_n_save_settings(folder:str, params:list[str] | None = None, extra_params=None, all_flag=False, save_file_name=None)-> tuple[dict , str]:

   # normalize the path delimiter
   folder = path.normpath(folder)
   
   # required params
   required_params = []
   if params is None:
      params = monthly_params.copy()
      
   required_params = params + extra_params if extra_params else params

   # read all contents of the folder
   files_in_dir = listdir(folder)

   # filter out unwanted files
   text_files = []

   for entry in files_in_dir:
      if path.isfile(path.join(folder, entry)) and entry.lower().endswith(".txt") :
         text_files.append(entry)

   # extract settings for each transmitter
   # master_param_dict = dict()

   # *** OPTIMIZATION: Use a list to store DataFrames, then concatenate once ***
   all_dictionaries = []

   # combine data from each file
   total = len(text_files)
   
   if total == 0:
      print("[WARN] No .txt files found to process.")
      return 
   
   print("[INFO] Processing started...\n")

   for i, text in enumerate(text_files, 1):
      print(f"[INFO] [{i}/{total}] Processing: {text}", flush=True)
      # Append the single-row DataFrame result to the list
      all_dictionaries.append(t6t_settings_parser( path.join(folder, text), required_params ))

   # Perform a single, much faster concatenation operation outside the loop
   master_param_dict = merge_dicts(all_dictionaries)

   # print(master_param_dict)
   # check if any save file name is provided# check if any save file name is provided
   if save_file_name:
      report_name = path.join(folder, save_file_name)
   else:
      # Fix: Extract only the folder name (e.g., 'FEB25') for the file_name
      folder_name = path.basename(folder)
      report_name = path.join(folder, folder_name + "_monthly.xlsx")

   dict_processed = master_param_dict if all_flag else {param:master_param_dict.get(param, []) for param in required_params}
   export_to_formatted_excel(dict_processed , name_check(report_name))

   # prints report details
   print("[INFO] Processing complete.\n")
   print(f"[INFO] Path processed: {folder}")
   print(f"[INFO_REPORT] REPORT SAVED AT: { report_name }")


   return dict_processed, report_name


## running code
# from sys import argv
# if len(argv):
if __name__ == "__main__":
    parameters = parse_cmd()
    get_n_save_settings(parameters.folder, parameters.params, parameters.extra_params, parameters.all, parameters.save_file_name)


